import math
import os
from typing import Dict, List, Any, Optional, Callable

import pandas as pd
import logging

LOGGER = logging.getLogger(__name__)


def _auto_fit_columns(ws, df: pd.DataFrame):
    # local import to avoid hard dependency at module import time
    try:
        from openpyxl.utils import get_column_letter
    except Exception as ex:
        raise ImportError("openpyxl is required for Excel export (auto-fit columns)") from ex
    for i, col in enumerate(df.columns, 1):
        max_length = max(
            [len(str(col))] + [len(str(cell)) for cell in df[col].fillna('').astype(str).tolist()]
        )
        adjusted_width = (max_length + 2)
        ws.column_dimensions[get_column_letter(i)].width = adjusted_width


def export_unique_records(
    df: pd.DataFrame,
    output_filepath: str,
    sheet_name: str = "Unique Candidates",
    progress_callback: Optional[Callable[[float, str], None]] = None,
) -> bool:
    """
    Export unique (non-duplicate) records to Excel.
    """
    # Do not overwrite existing files
    if os.path.exists(output_filepath):
        LOGGER.error('Export target already exists: %s', output_filepath)
        raise FileExistsError(f"File already exists: {output_filepath}")

    try:
        LOGGER.info('Exporting unique records to %s', output_filepath)
        try:
            import openpyxl  # quick check to provide clearer error if missing
        except Exception as ex:
            raise ImportError("openpyxl is required for exporting to Excel. Install with `pip install openpyxl`.") from ex

        if progress_callback:
            progress_callback(0.0, "Starting export of unique records")

        with pd.ExcelWriter(output_filepath, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name=sheet_name, index=False)
            workbook = writer.book
            worksheet = writer.sheets[sheet_name]

            # local imports for openpyxl styles
            from openpyxl.styles import Font, PatternFill

            # Format header
            header_font = Font(bold=True)
            header_fill = PatternFill(start_color='FFDCE6F1', end_color='FFDCE6F1', fill_type='solid')
            header_row = list(worksheet.iter_rows(min_row=1, max_row=1))
            if header_row:
                for cell in header_row[0]:
                    cell.font = header_font
                    cell.fill = header_fill

            # Auto-fit columns
            _auto_fit_columns(worksheet, df)

            # Freeze header row
            worksheet.freeze_panes = worksheet['A2']

        LOGGER.info('Exported unique records to %s', output_filepath)
        if progress_callback:
            progress_callback(1.0, f"Exported unique records to {output_filepath}")
        else:
            print(f"Exported unique records to {output_filepath}")
        return True
    except Exception as ex:
        LOGGER.exception('Export failed for unique records: %s', ex)
        if progress_callback:
            progress_callback(1.0, f"Export failed: {ex}")
        else:
            print(f"Export failed: {ex}")
        raise


def export_duplicate_records(
    duplicate_groups: Dict[str, List[int]],
    df: pd.DataFrame,
    output_filepath: str,
    sheet_name: str = "Duplicate Groups",
    progress_callback: Optional[Callable[[float, str], None]] = None,
) -> bool:
    """
    Export duplicate records grouped together with visual formatting.
    """
    # Do not overwrite existing files
    if os.path.exists(output_filepath):
        LOGGER.error('Export target already exists: %s', output_filepath)
        raise FileExistsError(f"File already exists: {output_filepath}")

    try:
        LOGGER.info('Exporting %d duplicate groups to %s', len(duplicate_groups), output_filepath)
        try:
            import openpyxl
        except Exception:
            raise ImportError("openpyxl is required for exporting to Excel. Install with `pip install openpyxl`.")

        if progress_callback:
            progress_callback(0.0, "Starting export of duplicate records")

        rows: List[Dict[str, Any]] = []
        total_groups = len(duplicate_groups)
        group_index = 0

        for gid, indices in duplicate_groups.items():
            group_index += 1
            # Determine matching key (phone/email) if present
            phones = df.loc[indices].get('normalized_phone') if 'normalized_phone' in df.columns else None
            emails = df.loc[indices].get('normalized_email') if 'normalized_email' in df.columns else None
            matching_key = ''
            key_type = 'unknown'
            if phones is not None:
                vals = [str(v) for v in phones.tolist() if v]
                if vals:
                    matching_key = max(set(vals), key=vals.count)
                    key_type = 'phone'
            if not matching_key and emails is not None:
                vals = [str(v) for v in emails.tolist() if v]
                if vals:
                    matching_key = max(set(vals), key=vals.count)
                    key_type = 'email'

            for idx in indices:
                row = df.loc[idx].to_dict()
                row['duplicate_group'] = gid
                row['matching_key'] = matching_key
                row['matching_key_type'] = key_type
                rows.append(row)

            if progress_callback:
                progress_callback(group_index / max(1, total_groups), f"Processed group {gid}")

        if not rows:
            # create empty dataframe
            out_df = pd.DataFrame(columns=list(df.columns) + ['duplicate_group', 'matching_key', 'matching_key_type'])
        else:
            out_df = pd.DataFrame(rows)

        with pd.ExcelWriter(output_filepath, engine='openpyxl') as writer:
            out_df.to_excel(writer, sheet_name=sheet_name, index=False)
            worksheet = writer.sheets[sheet_name]

            # local imports
            from openpyxl.styles import Font, PatternFill

            # Format header
            header_font = Font(bold=True)
            header_fill = PatternFill(start_color='FFF2F2F2', end_color='FFF2F2F2', fill_type='solid')
            header_row = list(worksheet.iter_rows(min_row=1, max_row=1))
            if header_row:
                for cell in header_row[0]:
                    cell.font = header_font
                    cell.fill = header_fill

            # Apply alternating fills per group to visually separate groups
            fills = [PatternFill(start_color='FFFFFFFF', end_color='FFFFFFFF', fill_type='solid'),
                     PatternFill(start_color='FFF7F7FF', end_color='FFF7F7FF', fill_type='solid')]

            current_row = 2  # 1-based Excel rows; header is row 1
            if not out_df.empty:
                last_gid = None
                fill_idx = 0
                for _, r in out_df.iterrows():
                    gid = r.get('duplicate_group')
                    if gid != last_gid:
                        fill_idx = (fill_idx + 1) % len(fills)
                        last_gid = gid
                    for col_idx in range(1, len(out_df.columns) + 1):
                        worksheet.cell(row=current_row, column=col_idx).fill = fills[fill_idx]
                    current_row += 1

            _auto_fit_columns(worksheet, out_df)
            worksheet.freeze_panes = worksheet['A2']

        LOGGER.info('Exported duplicate records to %s', output_filepath)
        if progress_callback:
            progress_callback(1.0, f"Exported duplicate records to {output_filepath}")
        else:
            print(f"Exported duplicate records to {output_filepath}")
        return True
    except Exception as ex:
        LOGGER.exception('Export failed for duplicate records: %s', ex)
        if progress_callback:
            progress_callback(1.0, f"Export failed: {ex}")
        else:
            print(f"Export failed: {ex}")
        raise
class ExportService:
    pass